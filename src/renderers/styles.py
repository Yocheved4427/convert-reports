"""
Shared Excel styling constants and helpers for renderers.
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ── Fonts ──────────────────────────────────────────────────────────────────────
HEADER_FONT = Font(name="Arial", size=14, bold=True)
SUBHEADER_FONT = Font(name="Arial", size=11, bold=True)
TABLE_HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
DATA_FONT = Font(name="Arial", size=10)
SUMMARY_FONT = Font(name="Arial", size=11, bold=True)

# ── Alignment (RTL) ───────────────────────────────────────────────────────────
RTL_CENTER = Alignment(horizontal="center", vertical="center",
                       wrap_text=True, readingOrder=2)
RTL_RIGHT = Alignment(horizontal="right", vertical="center",
                      readingOrder=2)
RTL_LEFT = Alignment(horizontal="left", vertical="center",
                     readingOrder=2)

# ── Fills ──────────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4",
                          fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3",
                           fill_type="solid")
SUMMARY_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA",
                           fill_type="solid")

# ── Borders ────────────────────────────────────────────────────────────────────
THIN_SIDE = Side(border_style="thin", color="000000")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE,
                     top=THIN_SIDE, bottom=THIN_SIDE)


def set_rtl_sheet(ws) -> None:
    """Configure a worksheet for right-to-left direction."""
    ws.sheet_view.rightToLeft = True
