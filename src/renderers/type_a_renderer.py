"""
Renderer for Type-A reports (detailed attendance with overtime).

Produces an Excel file mirroring the original PDF layout:
  • Title row with report name and month
  • Data table: Date | Day | Location | Entry | Exit | Break | Total | 100% | 125% | 150% | Notes
  • Summary row at bottom with totals
  • RTL direction, Hebrew column headers
"""

from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from src.models import TypeAReport
from src.renderers.base_renderer import BaseRenderer
from src.renderers.styles import (
    ALT_ROW_FILL, DATA_FONT, HEADER_FILL, HEADER_FONT, RTL_CENTER,
    SUBHEADER_FONT, SUMMARY_FILL, SUMMARY_FONT, TABLE_HEADER_FONT,
    THIN_BORDER, set_rtl_sheet,
)

logger = logging.getLogger(__name__)

# Column headers (Hebrew, RTL order → col A is rightmost visually)
_HEADERS = [
    "תאריך",       # Date
    "יום",          # Day
    "מקום עבודה",   # Location
    "כניסה",        # Entry
    "יציאה",        # Exit
    "הפסקה",        # Break
    'סה"כ שעות',    # Total hours
    "שעות 100%",    # Hours 100%
    "שעות 125%",    # Hours 125%
    "שעות 150%",    # Hours 150%
    "הערות",        # Notes
]

_COL_WIDTHS = [14, 10, 14, 10, 10, 10, 12, 12, 12, 12, 14]


class TypeARenderer(BaseRenderer):

    def render(self, report: TypeAReport, output_path: str | Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "דוח נוכחות מפורט"
        set_rtl_sheet(ws)

        row_num = 1

        # ── Title ──────────────────────────────────────────────────
        ws.merge_cells(start_row=row_num, start_column=1,
                       end_row=row_num, end_column=len(_HEADERS))
        title_cell = ws.cell(row=row_num, column=1,
                             value=f"דוח נוכחות מפורט – {report.month_year}")
        title_cell.font = HEADER_FONT
        title_cell.alignment = RTL_CENTER
        row_num += 2

        # ── Column headers ─────────────────────────────────────────
        for col_idx, header in enumerate(_HEADERS, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=header)
            cell.font = TABLE_HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = RTL_CENTER
            cell.border = THIN_BORDER
        row_num += 1

        # ── Data rows ──────────────────────────────────────────────
        for i, r in enumerate(report.rows):
            values = [
                r.date,
                r.day_of_week,
                r.location,
                r.entry_time,
                r.exit_time,
                r.break_minutes,
                r.total_hours,
                r.hours_100,
                r.hours_125,
                r.hours_150,
                r.notes,
            ]
            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.font = DATA_FONT
                cell.alignment = RTL_CENTER
                cell.border = THIN_BORDER
                if i % 2 == 1:
                    cell.fill = ALT_ROW_FILL
            row_num += 1

        # ── Summary row ────────────────────────────────────────────
        row_num += 1
        if report.summary:
            summary_label = ws.cell(row=row_num, column=1, value='סה"כ')
            summary_label.font = SUMMARY_FONT
            summary_label.fill = SUMMARY_FILL
            summary_label.alignment = RTL_CENTER
            summary_label.border = THIN_BORDER

            summary_data = [
                ("ימי עבודה", report.summary.work_days),
                ('סה"כ שעות', report.summary.total_hours),
                ("100%", report.summary.hours_100),
                ("125%", report.summary.hours_125),
                ("150%", report.summary.hours_150),
            ]
            # Place summary in a two-column layout below the table
            row_num += 1
            for label, val in summary_data:
                cell_l = ws.cell(row=row_num, column=1, value=label)
                cell_l.font = SUMMARY_FONT
                cell_l.fill = SUMMARY_FILL
                cell_l.alignment = RTL_CENTER
                cell_l.border = THIN_BORDER

                cell_v = ws.cell(row=row_num, column=2, value=val)
                cell_v.font = SUMMARY_FONT
                cell_v.fill = SUMMARY_FILL
                cell_v.alignment = RTL_CENTER
                cell_v.border = THIN_BORDER
                row_num += 1

        # ── Column widths ──────────────────────────────────────────
        for col_idx, width in enumerate(_COL_WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # ── Save ───────────────────────────────────────────────────
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        logger.info(f"Type-A rendered → {output_path}")
