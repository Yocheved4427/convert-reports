"""
Renderer for Type-N reports (simple monthly attendance with pay summary).

Produces an Excel file mirroring the original PDF layout:
  • Summary box at top: work days, total hours, hourly rate, total pay
  • Data table: Date | Day | Entry | Exit | Total
  • RTL direction, Hebrew column headers
"""

from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from src.models import TypeNReport
from src.renderers.base_renderer import BaseRenderer
from src.renderers.styles import (
    ALT_ROW_FILL, DATA_FONT, HEADER_FILL, HEADER_FONT, RTL_CENTER,
    SUBHEADER_FONT, SUMMARY_FILL, SUMMARY_FONT, TABLE_HEADER_FONT,
    THIN_BORDER, set_rtl_sheet,
)

logger = logging.getLogger(__name__)

# Column headers
_HEADERS = [
    "תאריך",       # Date
    "יום",          # Day
    "שעת כניסה",    # Entry time
    "שעת יציאה",    # Exit time
    'סה"כ שעות',    # Total hours
]

_COL_WIDTHS = [14, 12, 14, 14, 14]


class TypeNRenderer(BaseRenderer):

    def render(self, report: TypeNReport, output_path: str | Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "דוח נוכחות חודשי"
        set_rtl_sheet(ws)

        row_num = 1

        # ── Title ──────────────────────────────────────────────────
        ws.merge_cells(start_row=row_num, start_column=1,
                       end_row=row_num, end_column=len(_HEADERS))
        title_cell = ws.cell(row=row_num, column=1,
                             value=f"דוח נוכחות חודשי – {report.month_year}")
        title_cell.font = HEADER_FONT
        title_cell.alignment = RTL_CENTER
        row_num += 2

        # ── Summary box ────────────────────────────────────────────
        if report.summary:
            summary_items = [
                ("ימי עבודה בחודש", report.summary.work_days),
                ('סה"כ שעות חודשיות', report.summary.total_hours),
                ("מחיר לשעה", f"₪{report.summary.hourly_rate:.2f}"),
                ('סה"כ לתשלום', f"₪{report.summary.total_pay:.2f}"),
            ]
            for label, val in summary_items:
                cell_l = ws.cell(row=row_num, column=1, value=label)
                cell_l.font = SUMMARY_FONT
                cell_l.fill = SUMMARY_FILL
                cell_l.alignment = RTL_CENTER
                cell_l.border = THIN_BORDER

                cell_v = ws.cell(row=row_num, column=2, value=val)
                cell_v.font = SUMMARY_FONT
                cell_v.fill = SUMMARY_FILL
                cell_v.alignment = RTL_CENTER
                cell_v.border = THIN_BORDER
                row_num += 1
            row_num += 1

        # ── Column headers ─────────────────────────────────────────
        for col_idx, header in enumerate(_HEADERS, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=header)
            cell.font = TABLE_HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = RTL_CENTER
            cell.border = THIN_BORDER
        row_num += 1

        # ── Data rows ──────────────────────────────────────────────
        for i, r in enumerate(report.rows):
            values = [
                r.date,
                r.day_of_week,
                r.entry_time,
                r.exit_time,
                r.total_hours,
            ]
            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.font = DATA_FONT
                cell.alignment = RTL_CENTER
                cell.border = THIN_BORDER
                if i % 2 == 1:
                    cell.fill = ALT_ROW_FILL
            row_num += 1

        # ── Column widths ──────────────────────────────────────────
        for col_idx, width in enumerate(_COL_WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # ── Save ───────────────────────────────────────────────────
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        logger.info(f"Type-N rendered → {output_path}")
